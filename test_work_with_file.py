import os
from zipfile import ZipFile
from openpyxl import load_workbook
from pypdf import PdfReader
import pytest
import shutil

directory = os.path.join(os.path.dirname(__file__), "resources")
print(directory)


@pytest.fixture()
def test_create_zip_archive():
    if not os.path.exists(directory):
        os.mkdir(directory)

    with ZipFile("resources/test_archive.zip", "w") as myzip:  # создал архив
        myzip.write("test_file.txt")
        myzip.write("test_file.csv")
        myzip.write("test_file.xlsx")
        myzip.write("test_file.pdf")

    with ZipFile("resources/test_archive.zip") as myzip:
        print(myzip.namelist())  # отобразил содержимое архива
        myzip.extractall(path="resources")  # извлекаю все файлы из архива в папку resources
        shutil.rmtree('resources')
    #yield os.remove('resources/test_archive.zip')


def test_read_pdf(test_create_zip_archive):
    reader = PdfReader('resources/test_file.pdf')  # открываю файл
    text = (reader.pages[1].extract_text())  # извлёк текст со второй страницы
    assert 'Simple, Rapid, Effective, and Scalable' in text  # проверяю наличие текста на странице


def test_read_xlsx(test_create_zip_archive):
    workbook = load_workbook('resources/test_file.xlsx')
    sheet = workbook.active
    assert sheet.cell(row=2, column=2).value == 5  # проверяю что значение в указанной ячейке равно 5


def test_read_csv(test_create_zip_archive):
    read_csv = open('resources/test_file.csv')  # открыл файл
    content = read_csv.readlines()  # читаю текст файла
    rows = content[1]  # выбрал строку по которой буду делать проверку
    assert 'booker12;9012;Rachel;Booker' in rows


def test_read_csv(test_create_zip_archive):
    with open('resources/test_file.txt') as read_txt:
        content = read_txt.read()
        assert 'QA_GURU' in content
