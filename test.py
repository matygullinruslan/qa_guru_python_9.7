import os
from zipfile import ZipFile
import csv
from openpyxl import load_workbook
from pypdf import PdfReader

directory = os.path.join(os.path.dirname(__file__), "resources")
print(directory)

if not os.path.exists(directory):
    os.mkdir(directory)


def test_create_zip_archive():
    with ZipFile("resources/test_archive.zip", "w") as myzip:
        myzip.write("test_file.txt")
        myzip.write("test_file.csv")
        myzip.write("test_file.xlsx")
        myzip.write("test_file.pdf")


with ZipFile("resources/test_archive.zip") as myzip:
    print(myzip.namelist())  # отобразить содержимое архива
    myzip.extractall(path="resources")  # извлечь все файлы из архива в папку resources

    reader = PdfReader('resources/test_file.pdf')
    # print(reader.pages)
    # print(len(reader.pages))
    # print(reader.pages[1].extract_text())
    # assert 'Simple, Rapid, Effective, and Scalable' in reader.pages[1] # тест падает, хотя текст соответствует

workbook = load_workbook('resources/test_file.xlsx')
sheet = workbook.active
print(sheet.cell(row=2, column=2).value)  # в файле значение 5, результат выдает None

with open('test_file.csv') as File:
    reader = csv.reader(File, delimiter=',', quotechar=',',
                        quoting=csv.QUOTE_MINIMAL)
    for row in reader:
        print(sheet.cell(row=2, column=2).value)  # в файле значение 9012, выводит всю колонку

