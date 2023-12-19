import os
from zipfile import ZipFile
import sheet
from pypdf import PdfReader
from openpyxl import load_workbook

# directory = os.path.join(os.path.dirname(__file__), "resources")
# print(directory)


workbook = load_workbook('resources/test_file.xlsx')
workbook.active
print(sheet.cell(row=1, column=2).value)
